from telebot import types
import telebot
import openpyxl
import datetime

bot = telebot.TeleBot('API')


def get_rozklad():
    wb = openpyxl.load_workbook('rozklad.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    row = sheet.get_highest_row()
    column = sheet.get_highest_column()

    rz = []

    for i in range(1, row + 1):
        if sheet.cell(row=i, column=1).value == datetime.datetime(2019, 11, 11, 0, 0):
            for j in range(1, column + 1):
                rz.append(sheet.cell(row=i, column=j + 1).value)

    rozklad = rz[0:4]
    paru = f'1.{rz[0]}\n2.{rz[1]}\n3.{rz[2]}\n4.{rz[3]}'
    return paru


@bot.message_handler(commands=['start'])
def handle_start(message):
    msg = 'Привіт, я бот-розклад!!!!\nя покажу тобі розклад на пару днів\nжми на кнопку!!!!'
    keyboard = types.InlineKeyboardMarkup()
    button = types.InlineKeyboardButton(text="Розклад", callback_data='rozklad')
    keyboard.add(button)
    bot.send_message(message.chat.id, msg, reply_markup=keyboard)


@bot.message_handler(commands=['help'])
def handle_help(message):
    hlp = '''
    Для того щоб скористуватись ботом натисни /start
    І в тебе буде кнопка з розкладом.
    
    Розклад показує на 3 дні
    '''
    keyboard = types.InlineKeyboardMarkup()
    button = types.InlineKeyboardButton(text="Розклад", callback_data='test')
    keyboard.add(button)
    bot.send_message(message.chat.id, hlp, reply_markup=keyboard)


@bot.message_handler(func=lambda m: True)
def echo_text(message):
    bot.send_message(message.chat.id, 'ЙДИ НАХУЙ ТУТ ТІЛЬКИ РОЗКЛАД!!!!!!')


@bot.callback_query_handler(func=lambda call: True)
def callback_inline(call):
    # Если сообщение из чата с ботом
    if call.message:
        if call.data == "rozklad":
            bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text=str(get_rozklad()))


if __name__ == '__main__':
    bot.polling(none_stop=True)
